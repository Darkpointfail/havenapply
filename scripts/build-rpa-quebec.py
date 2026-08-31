#!/usr/bin/env python3
"""Build slim RPA Québec catalog from the complete verifiable workbook.

Usage: python3 scripts/build-rpa-quebec.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data/rpa/rpa_quebec_complet_verifiable_2025-12-31.xlsx"
# Fallback to legacy annuaire if the complete workbook is absent.
LEGACY = ROOT / "data/rpa/rpa_quebec_annuaire_2025-12-31.xlsx"
OUT = ROOT / "data/rpa/quebec-residences.json"

REGION_NAMES = {
    "1": "Bas-Saint-Laurent",
    "2": "Saguenay–Lac-Saint-Jean",
    "3": "Capitale-Nationale",
    "4": "Mauricie et Centre-du-Québec",
    "5": "Estrie",
    "7": "Outaouais",
    "8": "Abitibi-Témiscamingue",
    "9": "Côte-Nord",
    "10": "Nord-du-Québec",
    "11": "Gaspésie–Îles-de-la-Madeleine",
    "12": "Chaudière-Appalaches",
    "13": "Laval",
    "14": "Lanaudière",
    "15": "Laurentides",
    "16": "Montérégie",
    "61": "Montréal",
    "62": "Montréal",
    "63": "Montréal",
    "64": "Montréal",
    "65": "Montréal",
}

SERVICE_FLAGS = [
    ("repas", "Repas"),
    ("soins_hygiene", "Soins d'hygiène"),
    ("aide_alimentation", "Aide à l'alimentation"),
    ("bain", "Aide au bain"),
    ("habillage", "Habillage"),
    ("admin_des_medicaments", "Administration des médicaments"),
    ("assistance_a_la_mobilite", "Aide à la mobilité"),
    ("assistance_soins_non_reglementees", "Assistance aux soins"),
    ("soins_infirmiers", "Soins infirmiers"),
    ("entretien_menager", "Entretien ménager"),
    ("entretien_vetements_ou_literie", "Entretien des vêtements"),
    ("distribution_des_medicaments", "Distribution des médicaments"),
    ("loisirs", "Loisirs"),
]

SAFETY_FLAGS = [
    ("rampe_dacces", "Rampe d'accès"),
    ("gicleur", "Gicleurs"),
    ("generatrice", "Génératrice"),
    ("avertisseur_fumee", "Avertisseurs de fumée"),
    ("avertisseur_monox_carbone", "Avertisseurs de monoxyde"),
    ("detecteur_alarme_incendie", "Alarme incendie"),
    ("dispositif_securite_immeuble", "Dispositif de sécurité immeuble"),
]


def yes(v) -> bool:
    s = str(v or "").strip().upper()
    if s in {"OUI", "O", "YES", "Y", "1", "TRUE"}:
        return True
    if s.startswith("GICL"):  # Giclé
        return True
    return False


def num(v):
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".").strip()))
    except ValueError:
        return None


def slug(ref: str) -> str:
    clean = re.sub(r"[^a-zA-Z0-9]+", "-", str(ref).strip()).strip("-").lower()
    return f"rpa-{clean or 'unknown'}"


def cell(raw, idx, key):
    if key not in idx:
        return None
    return raw[idx[key]]


def main() -> None:
    path = XLSX if XLSX.exists() else LEGACY
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = "RPA_Consolidees" if "RPA_Consolidees" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    idx = {h: i for i, h in enumerate(headers)}

    # Support both categorie_rpa (legacy) and categories_rpa (complete workbook)
    cat_key = "categories_rpa" if "categories_rpa" in idx else "categorie_rpa"

    required = ["ref_registre", "nom_residence", "municipalite_rpa", "statut_residence"]
    for key in required:
        if key not in idx:
            raise SystemExit(f"Missing column: {key}")

    out = []
    skipped = 0
    for raw in rows_iter:
        statut = str(cell(raw, idx, "statut_residence") or "").strip()
        if statut != "Active":
            skipped += 1
            continue

        ref = str(cell(raw, idx, "ref_registre") or "").strip()
        name = str(cell(raw, idx, "nom_residence") or "").strip()
        if not ref or not name:
            skipped += 1
            continue

        region_code = str(cell(raw, idx, "region") or "").strip()
        services = []
        for col, label in SERVICE_FLAGS:
            if yes(cell(raw, idx, col)):
                services.append(label)

        if yes(cell(raw, idx, "clientele_risque_errance")):
            services.append("Clientèle à risque d'errance")

        safety = []
        for col, label in SAFETY_FLAGS:
            if yes(cell(raw, idx, col)):
                safety.append(label)

        postal = str(cell(raw, idx, "code_postal_rpa") or "").replace(" ", "").upper()
        if len(postal) == 6:
            postal = f"{postal[:3]} {postal[3:]}"

        phone = str(cell(raw, idx, "telephone_rpa") or "").strip() or None
        address = str(cell(raw, idx, "adresse_rpa") or "").strip()
        city = str(cell(raw, idx, "municipalite_rpa") or "").strip()
        category = str(cell(raw, idx, cat_key) or "").strip()
        cert = str(cell(raw, idx, "etat_certification") or "").strip()
        type_res = str(cell(raw, idx, "type_residence") or "").strip()
        security = str(cell(raw, idx, "securite") or "").strip() or None

        nurses_day = num(cell(raw, idx, "nbremplinfsemjour")) or 0
        nurses_eve = num(cell(raw, idx, "nbremplinfsemsoir")) or 0
        nurses_night = num(cell(raw, idx, "nbremplinfsemnuit")) or 0
        aides_day = num(cell(raw, idx, "nbremplassspabsemjour")) or 0

        record = {
            "id": slug(ref),
            "ref": ref,
            "name": name,
            "address": address,
            "city": city,
            "postal": postal or None,
            "phone": phone,
            "regionCode": region_code,
            "region": REGION_NAMES.get(region_code, f"Région {region_code}"),
            "mrc": str(cell(raw, idx, "mrc") or "").strip() or None,
            "category": category,
            "type": type_res or None,
            "certification": cert or None,
            "capacity": num(cell(raw, idx, "capacite_daccueil_des_personnes_dans_la_rpa")),
            "residents": num(cell(raw, idx, "total_de_residents_rpa")),
            "units": num(cell(raw, idx, "total_unite_rpa")),
            "unitsByCategory": {
                "1": num(cell(raw, idx, "unites_rpa_categorie_1")),
                "2": num(cell(raw, idx, "unites_rpa_categorie_2")),
                "3": num(cell(raw, idx, "unites_rpa_categorie_3")),
                "4": num(cell(raw, idx, "unites_rpa_categorie_4")),
            },
            "roomsSingle": num(cell(raw, idx, "chambres_simples_rpa")),
            "roomsDouble": num(cell(raw, idx, "chambres_doubles_rpa")),
            "apartments": num(cell(raw, idx, "logements_rpa")),
            "services": services,
            "safety": safety,
            "security": security,
            "floors": num(cell(raw, idx, "nbr_detages_dans_limmeuble")),
            "elevators": num(cell(raw, idx, "nbr_dascenseurs_reguliers")),
            "openedOn": str(cell(raw, idx, "date_ouverture") or "").strip() or None,
            "ages": {
                "under65": num(cell(raw, idx, "age_65")),
                "from65to74": num(cell(raw, idx, "age_65_a_74")),
                "from75to84": num(cell(raw, idx, "age_75_a_84")),
                "from85": num(cell(raw, idx, "age_85_et")),
            },
            "staffing": {
                "nursesWeekday": {"day": nurses_day, "evening": nurses_eve, "night": nurses_night},
                "aidesWeekdayDay": aides_day,
                "hasNursingPresence": bool(nurses_day or nurses_eve or nurses_night),
            },
            "entente108": yes(cell(raw, idx, "unite_locative_rpa_avec_entente_108")),
            "operator": str(cell(raw, idx, "nom_de_lentreprise") or "").strip() or None,
            "group": str(cell(raw, idx, "regroupement") or "").strip() or None,
            "verified": str(cell(raw, idx, "statut_verification") or "").strip() or None,
            "sourceDate": str(cell(raw, idx, "date_extraction") or "").strip() or None,
        }
        out.append(record)

    out.sort(key=lambda r: (r["region"], r["city"], r["name"]))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": "Registre des RPA — Québec (complet vérifiable)",
        "extractedOn": "2025-12-31",
        "workbook": path.name,
        "count": len(out),
        "residences": out,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(out)} active residences → {OUT.relative_to(ROOT)} from {path.name} (skipped {skipped})")


if __name__ == "__main__":
    main()
